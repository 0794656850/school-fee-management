from pathlib import Path
import re

app_path = Path('app.py')
src = app_path.read_text(encoding='utf-8')

# 1) Add reports() route if missing
if 'def reports()' not in src:
    insert_pt = src.find('\n# ---------- RUN ----------')
    block = """

# ---------- REPORTS PAGE ----------
@app.route("/reports")
def reports():
    return render_template("reports.html")
"""
    if insert_pt != -1:
        src = src[:insert_pt] + block + src[insert_pt:]
    else:
        src += block

# 2) Add export_fees_xlsx route if missing
if 'def export_fees_xlsx()' not in src:
    insert_pt = src.find('\n# ---------- RUN ----------')
    block = """

# ---------- XLSX EXPORT (optional) ----------
@app.route("/export_fees_xlsx")
def export_fees_xlsx():
    if not is_pro_enabled(app):
        try:
            flash('Exports are available in Pro. Please upgrade.', 'warning')
        except Exception:
            pass
        return redirect(url_for('monetization.index'))
    try:
        from openpyxl.workbook import Workbook
    except Exception:
        try:
            flash('XLSX export requires openpyxl. Falling back to ZIP of CSVs.', 'warning')
        except Exception:
            pass
        return redirect(url_for('export_fees_full'))

    db = get_db_connection()
    cur = db.cursor(dictionary=True)
    sid = session.get('school_id')

    cur.execute("SELECT name, admission_no, class_name, COALESCE(balance, fee_balance) AS balance, COALESCE(credit,0) AS credit FROM students WHERE school_id=%s ORDER BY class_name, name", (sid,))
    students = cur.fetchall() or []

    cur.execute("SELECT s.name, s.admission_no, s.class_name, p.year, p.term, p.amount, p.method, p.reference, p.date FROM payments p JOIN students s ON s.id=p.student_id WHERE p.school_id=%s ORDER BY p.date DESC", (sid,))
    payments = cur.fetchall() or []

    cur.execute("SELECT class_name AS class, COUNT(*) AS total_students, COALESCE(SUM(COALESCE(balance,fee_balance)),0) AS total_pending, COALESCE(SUM(credit),0) AS total_credit FROM students WHERE school_id=%s GROUP BY class_name ORDER BY class_name", (sid,))
    class_summary = cur.fetchall() or []

    cur.execute("SELECT p.year AS year, p.term AS term, COALESCE(SUM(p.amount),0) AS total FROM payments p WHERE p.school_id=%s AND p.method <> 'Credit Transfer' GROUP BY p.year, p.term ORDER BY p.year, p.term", (sid,))
    term_summary = cur.fetchall() or []

    cur.execute("SELECT p.method AS method, COUNT(*) AS cnt, COALESCE(SUM(p.amount),0) AS total FROM payments p WHERE p.school_id=%s AND p.method <> 'Credit Transfer' GROUP BY p.method ORDER BY total DESC", (sid,))
    method_breakdown = cur.fetchall() or []

    db.close()

    from io import BytesIO
    from openpyxl.workbook import Workbook
    wb = Workbook()
    ws = wb.active; ws.title = 'Students'
    ws.append(['Name','Admission No','Class','Balance (KES)','Credit (KES)'])
    for r in students:
        ws.append([r.get('name'), r.get('admission_no'), r.get('class_name'), r.get('balance'), r.get('credit')])

    ws2 = wb.create_sheet('Payments')
    ws2.append(['Student Name','Admission No','Class','Year','Term','Amount (KES)','Method','Reference','Date'])
    for p in payments:
        ws2.append([p.get('name'), p.get('admission_no'), p.get('class_name'), p.get('year'), p.get('term'), p.get('amount'), p.get('method'), p.get('reference'), p.get('date')])

    ws3 = wb.create_sheet('Class Summary')
    ws3.append(['Class','Total Students','Total Pending (KES)','Total Credit (KES)'])
    for c in class_summary:
        ws3.append([c.get('class'), c.get('total_students'), c.get('total_pending'), c.get('total_credit')])

    ws4 = wb.create_sheet('Term Summary')
    ws4.append(['Year','Term','Total Collected (KES)'])
    for t in term_summary:
        ws4.append([t.get('year'), t.get('term'), t.get('total')])

    ws5 = wb.create_sheet('Method Breakdown')
    ws5.append(['Method','Count','Total (KES)'])
    for m in method_breakdown:
        ws5.append([m.get('method'), m.get('cnt'), m.get('total')])

    mem = BytesIO(); wb.save(mem); mem.seek(0)
    from datetime import datetime as _dt
    ts = _dt.now().strftime('%Y%m%d_%H%M%S')
    return Response(mem.getvalue(), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={'Content-Disposition': f'attachment; filename=fees_full_report_{ts}.xlsx'})
"""
    if insert_pt != -1:
        src = src[:insert_pt] + block + src[insert_pt:]
    else:
        src += block

app_path.write_text(src, encoding='utf-8')
print('app.py updated with /reports and /export_fees_xlsx')
